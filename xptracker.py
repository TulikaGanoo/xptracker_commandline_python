#defining all variables (since some of them are 'hidden' in functions)
learn_new = ""
learn_new_items = ""
learn_new_list = ""
number_of_learn_new = 0
power = 0
total_power = 0
group_standing_xp = 0
fun_xp = 0
foe_list = ""
foe_list_items =""
current_level = 0


#Setting up - opening excel, getting current xp
import openpyxl
import pandas as pd
from openpyxl import load_workbook


path = 'XpTrackerData.xlsx'
wb = load_workbook(path)
sheet = wb.active
max_row=sheet.max_row

current_level = sheet.cell(max_row, column = 8).value
current_xp = sheet.cell(max_row, column = 9).value

print(f"current level {current_level}, current xp {current_xp}")

#Getting today's date
from datetime import datetime

now = datetime.now()
dt_string = now.strftime("%d/%m/%Y %H:%M:%S")




#asking questions, getting answers

check_for_affirmative = ["yes", "yeah", "yup", "yeahboy"]

learn_new = input("Did you learn something new today? ")
for x in check_for_affirmative:
    if learn_new.count(x) > 0:
        learn_new_items = input("Let's hear 'em! Separate each new thing learnt with a comma to give em some space ")
        learn_new_list = learn_new_items.split(",")
        number_of_learn_new =len(learn_new_list)

foe = input("Did you defeat a forbiddable foe? ")
for x in check_for_affirmative:
    if foe.count(x) > 0:
        foe_list_items = input("Let's hear 'em! Separate each foe with a comma to give em some space ")
        foe_list = foe_list_items.split(",")
        number_of_foes = len(foe_list)
        for x in range (0,number_of_foes):
            power = int(input(f"How powerful was foe {foe_list[x]} where 0 = no foe, 10 = most powerful foe "))
            total_power =+ power

groupstanding = input("Did you improve your group standing today? ")

fun = input("Did you have fun? ")

#calculating xp
learn_new_xp = number_of_learn_new * 10
foe_xp = total_power * 100;
for x in check_for_affirmative:
    if groupstanding.count(x) > 0:
        group_standing_xp = 100
for x in check_for_affirmative:
    if fun.count(x) > 0:
        fun_xp = 100

total_gained_xp = int(learn_new_xp) + int(foe_xp) + int(group_standing_xp) + int(fun_xp)
print("Great work today!")
print(f"You had {current_xp} xp")
print(f"You gained {total_gained_xp} xp")

current_xp += total_gained_xp
print(f"You now have {current_xp} xp")

print(f"you were at level {current_level}")

#getting levels
level_list = [0,300,900,2700,6500,14000,23000,34000,48000,64000,85000,100000,120000,140000,165000,195000,225000,265000,305000,355000]
for level, level_threshold in enumerate(level_list):
    if current_xp >= level_threshold:
        current_level = level + 1
#printing out xp

print(f"You now have {current_xp} xp and are at level {current_level}")


#Saving all info to our excel data sheet
inputdata = [dt_string, learn_new_items, foe_list_items, total_power, groupstanding, fun, total_gained_xp, current_level, current_xp]

sheet.append(inputdata)

wb.save(path)


